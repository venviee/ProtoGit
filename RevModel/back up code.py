from django.shortcuts import render, redirect
import openpyxl
import shutil
import os
import json

# Determine the directory of the current file (views.py) and build the path to config.json
current_file_dir = os.path.dirname(__file__)
config_path = os.path.join(current_file_dir, 'config.json')

with open(config_path) as config_file:
    config = json.load(config_file)

# Helper function to parse staged growth inputs
def parse_staged_input(form_list_years, form_list_rates):
    return [(int(year), float(rate)) for year, rate in zip(form_list_years, form_list_rates)]

# Helper function to delete columns beyond a specified index in an Excel sheet
def delete_columns_beyond_index(ws, end_col_index):
    max_col = ws.max_column
    for col in range(max_col, end_col_index, -1):
        ws.delete_cols(col)

# Function to update growth rates in an Excel sheet
def update_growth_rates(ws, row, start_col_index, number_of_years, growth_choice, constant_rate=None, stages=None):
    end_col_index = start_col_index + number_of_years - 1
    if growth_choice == "constant" and constant_rate is not None:
        growth_rate = constant_rate / 100
        for col in range(start_col_index, end_col_index + 1):
            cell = ws.cell(row=row, column=col)
            cell.value = growth_rate
            cell.number_format = '0.00%'
    elif growth_choice == "staged" and stages:
        current_col_index = start_col_index
        for duration, rate in stages:
            growth_rate = rate / 100
            for _ in range(duration):
                if current_col_index > end_col_index:
                    break
                cell = ws.cell(row=row, column=current_col_index)
                cell.value = growth_rate
                cell.number_format = '0.00%'
                current_col_index += 1

# Function for handling revenue questions form
def revenue_questions(request):
    if request.method == 'POST':
        # Use getlist to ensure data is saved as lists
        revenue_data = {
            'number_of_years': request.POST.get('number_of_years'),
            'revenue_choice': request.POST.get('revenue_choice'),
            'revenue_constant_growth_rate': request.POST.get('revenue_constant_growth_rate', '0'),
            'revenueStageYears': request.POST.getlist('revenueStageYears[]'),
            'revenueStageGrowthRates': request.POST.getlist('revenueStageGrowthRates[]'),
        }
        request.session['revenue_data'] = revenue_data
        request.session.modified = True  # Make sure the session is saved
        return redirect('expense_questions')
    else:
        # If we're not posting, we're loading the form
        # Load data from session if it exists to pre-populate the form
        revenue_data = request.session.get('revenue_data', {})
        # Pass revenue_data to the context to pre-populate the form fields
    return render(request, 'revenue_questions.html', {'revenue_data': revenue_data})

def grossmargin_questions(request):
    if request.method == 'POST':
        # Use getlist to ensure data is saved as lists
        grossmargin_data = {
            'number_of_years': request.POST.get('number_of_years'),
            'grossmargin_choice': request.POST.get('grossmargin_choice'),
            'grossmargin_constant_growth_rate': request.POST.get('grossmargin_constant_growth_rate', '0'),
            'grossmarginStageYears': request.POST.getlist('grossmarginStageYears[]'),
            'grossmarginStageGrowthRates': request.POST.getlist('grossmarginStageGrowthRates[]'),
        }
        request.session['grossmargin_data'] = grossmargin_data
        request.session.modified = True  # Make sure the session is saved
        return redirect('expense_questions')
    else:
        # If we're not posting, we're loading the form
        # Load data from session if it exists to pre-populate the form
        grossmargin_data = request.session.get('grossmargin_data', {})
        # Pass revenue_data to the context to pre-populate the form fields
    return render(request, 'grossmargin_questions.html', {'grossmargin_data': grossmargin_data})

def expense_questions(request):
    if request.method == 'POST':
        expense_data = {
            'expense_choice': request.POST.get('expense_choice', request.session.get('expense_data', {}).get('expense_choice', 'constant')),
            'expense_constant_growth_rate': request.POST.get('expense_constant_growth_rate', request.session.get('expense_data', {}).get('expense_constant_growth_rate', '0')),
            'expenseStageYears': request.POST.getlist('expenseStageYears[]') or request.session.get('expense_data', {}).get('expenseStageYears', []),
            'expenseStageGrowthRates': request.POST.getlist('expenseStageGrowthRates[]') or request.session.get('expense_data', {}).get('expenseStageGrowthRates', []),
        }


        # Accessing config values
        original_path = config['original_path']
        file_name = config['file_name']
        copy_name = config['copy_name_prefix'] + file_name
        start_column_index = config['start_column_index']

        request.session['expense_data'] = expense_data  # Update session with current expense data
        request.session.modified = True

        # Process both sets of data and update Excel
        shutil.copy(original_path, copy_name)

        # Ensure the directory exists
        os.makedirs(os.path.dirname(copy_name), exist_ok=True)

        wb = openpyxl.load_workbook(copy_name)
        ws_inputs = wb[config['worksheet_names']['inputs']]

        revenue_data = request.session.get('revenue_data')
        number_of_years = int(revenue_data.get('number_of_years', 0))
        start_column_index = config['start_column_index']

        # Process revenue data
        update_growth_rates(ws_inputs, config['assigned_rows']['revenue'], start_column_index, number_of_years, revenue_data.get('revenue_choice'), constant_rate=float(revenue_data.get('revenue_constant_growth_rate', '0') or 0), stages=parse_staged_input(revenue_data.get('revenueStageYears', []), revenue_data.get('revenueStageGrowthRates', [])))
        
  
        # Process expense data
        update_growth_rates(ws_inputs, config['assigned_rows']['expense'], start_column_index, number_of_years, expense_data.get('expense_choice'), constant_rate=float(expense_data.get('expense_constant_growth_rate', '0') or 0), stages=parse_staged_input(expense_data.get('expenseStageYears', []), expense_data.get('expenseStageGrowthRates', [])))

        delete_columns_beyond_index(ws_inputs, start_column_index + number_of_years - 1)
        delete_columns_beyond_index(wb[config['worksheet_names']['model']], start_column_index + number_of_years - 1)

        wb.save(copy_name)
        return render(request, 'result.html')
    elif 'revenue_data' not in request.session:
        return redirect('revenue_questions')
    else:
        revenue_data = request.session.get('revenue_data', {})
        return render(request, 'expense_questions.html', {'number_of_years': revenue_data.get('number_of_years', 0)})

